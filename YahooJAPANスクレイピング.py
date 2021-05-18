#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Thu Dec 31 22:45:25 2020

@author: macbookleo
"""

import requests
from bs4 import BeautifulSoup
import openpyxl as px
import urllib

def search():
    global start, end, l
    x = int(end) - int(start) + 1
    for i in range(x):
        cell = ws["B"+str(start)].value
        load_url= "http://search.yahoo.co.jp/search?p="+urllib.parse.quote_plus(cell)+" 企業情報 概要"
        html = requests.get(load_url)
        soup = BeautifulSoup(html.text, "lxml")
        for link in soup.find_all("a"):
            l.append(link["href"])
        ws["U" + str(start)] = l[2]
        print(l[2])
        l=[]
        start = int(start) + 1
    
def save():
    wb.save('ユーザー属性調査(企業業態・職種）.xlsx')
if __name__=="__main__":
    l=[]
    wb = px.load_workbook('ユーザー属性調査(企業業態・職種）.xlsx')
    ws = wb['For check']
    start=int(input("始め:"))
    end=int(input("終わり:"))
    search()
    save()