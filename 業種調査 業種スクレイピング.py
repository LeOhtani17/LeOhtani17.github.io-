# -*- coding: utf-8 -*-
"""
Spyder Editor

This is a temporary script file.
"""

from bs4 import BeautifulSoup
import openpyxl as px
import requests

def search1():
    global URL
    html = requests.get(URL)
    soup = BeautifulSoup(html.content,'html.parser')
    text=soup.find_all("div",class_="text")
    for i in range(2,15714):
        cell = ws["B"+str(i)].value
        if "株式会社" in str(cell):
            cell=cell.strip("株式会社")
            if cell in str(text):
                ws["S"+str(i)]="成形加工メーカー"
def search2():
    global sort
    for i in range(2,15714):
        cell = ws["B"+str(i)].value
        if sort in str(cell):
            ws["S"+str(i)]="製品メーカー"
def search3():
    global count, l
    while True:
        url="https://baseconnect.in/companies/category/767226a3-66d7-4af9-a70c-0c4af552ac4a?page="+str(count)
        html = requests.get(url)
        soup = BeautifulSoup(html.content,'html.parser')
        text=soup.find_all("h4",class_="searches__result__list__header__title")
        l.append(text)
        count=count+1
        if "次へ" not in str(soup):
            break
    for i in range(2,15714):
        cell = ws["B"+str(i)].value
        if "株式会社" in str(cell):
            cell=cell.strip("株式会社")
            cell=cell.strip(" 株式会社")
            if str(cell) in str(l):
                ws["S"+str(i)]="金型メーカー"
        elif str(cell) in str(l):
            ws["S"+str(i)]="金型メーカー"
    
def save():
    wb.save('ユーザー属性調査(企業業態・職種）.xlsx')
if __name__=="__main__":
    wb = px.load_workbook('ユーザー属性調査(企業業態・職種）.xlsx')
    ws = wb['For check']
    sort="製作所"
    l=[]
    count=1
    URL="https://rnavi.ndl.go.jp/mokuji_html/027595729-01.html"
    #search1()
    #search2()
    search3()
    save()