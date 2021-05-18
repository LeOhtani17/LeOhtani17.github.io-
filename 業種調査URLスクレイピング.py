#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Mon Dec 21 15:40:57 2020

@author: macbookleo
"""

from bs4 import BeautifulSoup
import openpyxl as px
import requests
import time

def search():
    global start, end
    x = int(end) - int(start) + 1
    for i in range(x):
        cell = ws["B"+str(start)].value
        URL= "http://search.yahoo.com/search?p="+str(cell)+" 企業情報 概要"
        raw_page = requests.get(URL)
        soup = BeautifulSoup(raw_page.text,'lxml')
        for link in soup.select(".ac-algo.fz-l.ac-21th.lh-24"):
            ws["U" + str(start)] = link.get("href")
            print(link.get("href"))
            break
        start = int(start) + 1
        time.sleep(2)
    
def save():
    wb.save('ユーザー属性調査(企業業態・職種）.xlsx')
if __name__=="__main__":
    wb = px.load_workbook('ユーザー属性調査(企業業態・職種）.xlsx')
    ws = wb['For check']
    start=int(input("始め:"))
    end=int(input("終わり:"))
    search()
    save()