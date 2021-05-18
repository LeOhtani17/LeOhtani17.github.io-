#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Sun Oct 11 22:13:57 2020

@author: reo
"""

import  tkinter as tk
import openpyxl as px
import requests
import urllib
from bs4 import BeautifulSoup

def search():
    global start, end, l
    start_num = start.get()
    end_num = end.get()
    x = int(end_num) - int(start_num) + 1
    for i in range(x):
        cell = ws["B"+str(start_num)].value
        URL= "http://search.yahoo.com/search?p="+urllib.parse.quote_plus(cell, encoding='utf-8')+" 企業情報"
        raw_page = requests.get(URL)
        soup = BeautifulSoup(raw_page.text,'lxml')
        for link in soup.select(".ac-algo.fz-l.ac-21th.lh-24"):
            ws["U" + str(start_num)] = link['href']
            break
        load_url = link['href']
        html = requests.get(load_url)
        soup_2 = BeautifulSoup(html.content, "html.parser")

        for j in soup_2.find_all(string=('企業情報','会社案内','会社概要','企業概要')):
            l.append(j.parent)
            url="'''"+str(l)+"'''"
            soup_3 = BeautifulSoup(url, 'html.parser')
            l=[]
            if soup_3 == None:
                ws["Z" + str(start_num)]= " "
            else:
                for link1 in soup_3.find_all('a'):
                    ws["Z" + str(start_num)]=link1.get("href")
                
        start_num = int(start_num) + 1

def save():
    wb.save('ユーザー属性調査(企業業態・職種）.xlsx')

if __name__=="__main__":
    win = tk.Tk()
    win.geometry("300x300")
    win.title('URL-Excel')
    cvs = tk.Canvas(win, width=300, height=300)
    cvs.pack()
    #cvs.create_rectangle(0, 0, 300, 300, fill = 'white')

    wb = px.load_workbook('ユーザー属性調査(企業業態・職種）.xlsx')
    ws = wb['For check']

    start = tk.Entry(width=8)
    start.place(x=120, y=40)
    end = tk.Entry(width=8)
    end.place(x=120, y=70)
    label1 = tk.Label(win, text='start', fg='green', font=('Helvetica', 15))
    label1.place(x=80, y=40)
    label2 = tk.Label(win, text='end', fg='green', font=('Helvetica', 15))
    label2.place(x=80, y=70)
    
    btn_s = tk.Button(win, text=u'search', command=search, width=15)
    btn_s.place(x=55, y=120)
    btn_save = tk.Button(win, text=u'save Excel', command=save, width=15)
    btn_save.place(x=55, y=200)
    y = 0
    l=[]

win.mainloop()

