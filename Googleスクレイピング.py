#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Thu Dec 31 15:07:43 2020

@author: macbookleo
"""

import openpyxl as px
import requests
import urllib
from bs4 import BeautifulSoup
import re

def search():
    global start, end, l, cleansing
    l=[]
    x = int(end) - int(start) + 1
    for i in range(x):
        cell = ws["B"+str(start)].value
        URL= "https://www.google.com/search?q="+urllib.parse.quote_plus(str(cell))+" 企業情報 概要"
        raw_page = requests.get(URL)
        soup = BeautifulSoup(raw_page.text,'lxml')
        for link in soup.find_all('a'):
            l.append(link['href'])
        cleansing=re.search('(?<==).*(?=&sa)', l[17]).group()
        print(cleansing)
        ws["U" + str(start)] = str(cleansing)
        l=[]
        start = int(start) + 1
    
def save():
    wb.save('ユーザー属性調査(企業業態・職種）.xlsx')
if __name__=="__main__":
    wb = px.load_workbook('ユーザー属性調査(企業業態・職種）.xlsx')
    ws = wb['For check']
    start=int(input("始め:"))
    end=int(input("終わり:"))
    search()
    save()