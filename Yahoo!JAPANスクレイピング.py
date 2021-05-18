#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue Jan  5 02:10:54 2021

@author: macbookleo
"""

import requests
from requests import Response
from bs4 import BeautifulSoup
import openpyxl as px

def main():
    global start, end
    x = int(end) - int(start) + 1
    for i in range(x):
        cell = ws["B"+str(start)].value
        URL= "http://search.yahoo.co.jp/search?p="+str(cell)+" 企業情報 概要"
        response: Response = requests.get(URL)
        soup: BeautifulSoup = BeautifulSoup(response.text,'lxml')
        ws["U" + str(start)] = soup.body.footer.div.p.a["href"]
        print(soup.body.footer.div.p.a["href"])
        start = int(start) + 1


def save():
    wb.save('ユーザー属性調査(企業業態・職種）.xlsx')

if __name__ == "__main__":
    wb = px.load_workbook('ユーザー属性調査(企業業態・職種）.xlsx')
    ws = wb['For check']
    start=int(input("始め:"))
    end=int(input("終わり:"))
    main()
